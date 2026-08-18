"""Tests for export.py: build_skeleton_rows (the chronological hotel/flight
merge shared with the /preview/data endpoint) and its use inside build_excel.
"""

from openpyxl import load_workbook

import export

TRIP_DATA_WITH_SKELETON = {
    "recommendations": [],
    "hotels": [
        {
            "name": "Gracery Shinjuku",
            "city": "Tokyo",
            "check_in": "2026-09-04",
            "check_out": "2026-09-11",
        },
    ],
    "flights": [
        {
            "from": "TLV",
            "to": "NRT",
            "description": "Outbound",
            "departure": "2026-09-03 23:00",
            "arrival": "2026-09-04 18:00",
        },
    ],
}


def test_build_skeleton_rows_sorts_flights_and_hotels_chronologically():
    rows = export.build_skeleton_rows(TRIP_DATA_WITH_SKELETON)

    assert [row["type"] for row in rows] == ["Flight", "Hotel"]
    assert rows[0]["date"] == "2026-09-03"
    assert rows[0]["location_route"] == "TLV -> NRT"
    assert rows[1]["date"] == "2026-09-04"
    assert rows[1]["location_route"] == "Tokyo"


def test_build_excel_skeleton_sheet_matches_build_skeleton_rows():
    rows = export.build_skeleton_rows(TRIP_DATA_WITH_SKELETON)

    workbook = load_workbook(export.build_excel(TRIP_DATA_WITH_SKELETON))
    sheet = workbook["Trip Skeleton"]

    assert [cell.value for cell in sheet[1]] == ["Date", "Type", "Location/Route", "Details"]
    for row_cells, expected in zip(sheet.iter_rows(min_row=2), rows):
        assert [cell.value for cell in row_cells] == [
            expected["date"], expected["type"], expected["location_route"], expected["details"]
        ]
